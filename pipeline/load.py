#!/usr/bin/env python3
"""load.py — serialize cleaned data + reliability report into a self-certifying workbook.

Release gate (physical, not an approval step): emit `<base>_RELEASE.xlsx` iff zero HARD checks fail
after reliability AND nothing was quarantined; otherwise `<base>_DRAFT.xlsx` with quarantined values
already nulled. Every workbook carries VALIDATION, QUARANTINE, RELIABILITY sheets and a README A1
STATUS stamp.

Output is byte-deterministic: doc timestamps are fixed and the xlsx zip is normalized (sorted
entries, fixed date_time), so two runs on the same input produce identical bytes.

`extract(data, out_dir, ...)` runs reliability then writes the workbook.
"""
import os
import zipfile
from datetime import datetime
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from . import reliability as REL
from . import validate as V

_FIXED_DT = datetime(2020, 1, 1, 0, 0, 0)
_BOLD = Font(bold=True)

TITLES = {"constructs": "0_Construct", "review": "1_Review", "study": "2_Study", "arms": "3_Arm",
          "baseline": "4_BaselineChar", "outcomes": "5_Outcome", "comparisons": "6_Comparison",
          "armdata": "7_ArmData", "estimates": "8_Estimate", "rob": "9_RoB"}
ORDER = {
    "estimates": ["estimate_id", "outcome_id", "comparison_id", "effect_measure", "point_estimate",
                  "ci_lower", "ci_upper", "ci_level", "se", "p_value", "p_operator", "p_sided"],
    "armdata": ["data_id", "arm_id", "outcome_id", "value_basis", "n", "mean", "sd", "sem",
                "median", "events", "total", "pct"],
    "baseline": ["study_id", "arm_id", "characteristic", "value", "n", "dispersion_type",
                 "dispersion_value", "dispersion_low", "dispersion_high"],
}
PROV = ["source_modality", "derivation_method", "confidence", "extractor"]


def _rows(data, key):
    v = data.get(key)
    return [] if v is None else (v if isinstance(v, list) else [v])


def _columns(sheet_key, recs):
    keys = []
    for pref in ORDER.get(sheet_key, []):
        if any(pref in r for r in recs) and pref not in keys:
            keys.append(pref)
    extras = sorted({k for r in recs for k in r if not k.startswith("_")} - set(keys) - set(PROV))
    keys.extend(extras)
    keys.extend([p for p in PROV if any(p in r for r in recs)])
    return keys


def _write_sheet(wb, title, recs):
    ws = wb.create_sheet(title)
    if not recs:
        return
    cols = _columns_from_key(title, recs)
    for j, c in enumerate(cols, 1):
        ws.cell(1, j, c).font = _BOLD
    for i, r in enumerate(recs, 2):
        for j, c in enumerate(cols, 1):
            v = r.get(c, "")
            ws.cell(i, j, "" if v is None else v)


def _columns_from_key(title, recs):
    key = next((k for k, t in TITLES.items() if t == title), title)
    return _columns(key, recs)


def build_workbook(clean_data: Dict, report: Dict, out_dir: str, base_name: str = "extraction",
                   alpha: float = 0.05) -> Dict:
    run_ts = report.get("run_ts") or "1970-01-01T00:00:00+00:00"
    checks = report.get("checks") or V.validate_data(clean_data, alpha)["checks"]
    hard_fail = sum(c["fail"] for c in checks if c["status"] == "FAIL")
    q = report.get("quarantine") or []
    release = hard_fail == 0 and len(q) == 0
    status = (f"STATUS: RELEASE — all hard checks pass, 0 withheld  @ {run_ts}" if release else
              f"STATUS: DRAFT — NOT RELEASE READY ({len(q)} quarantined / {hard_fail} hard-fail)  @ {run_ts}")

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.created = _FIXED_DT
    wb.properties.modified = _FIXED_DT

    # README (A1 = status)
    rd = wb.create_sheet("README")
    lines = [status, "", "ExtractApp v3A — autonomous reliability layer.",
             "Quarantined numbers are NULLED in the data sheets and listed in QUARANTINE.",
             "VALIDATION = the hard/soft/guarded battery over the cleaned data.",
             "RELIABILITY = per-run telemetry.", ""]
    tel = report.get("telemetry") or {}
    lines.append(f"numbers={tel.get('numbers', 0)} confirmed={tel.get('confirmed', 0)} "
                 f"quarantined={tel.get('quarantined', 0)} needs_vision={tel.get('needs_vision', 0)}")
    for i, ln in enumerate(lines, 1):
        rd.cell(i, 1, ln)

    # data sheets in canonical order
    for key in TITLES:
        recs = _rows(clean_data, key)
        if recs:
            _write_sheet(wb, TITLES[key], recs)

    _validation_sheet(wb, checks, run_ts)
    _quarantine_sheet(wb, q)
    _reliability_sheet(wb, tel, run_ts)

    os.makedirs(out_dir, exist_ok=True)
    suffix = "RELEASE" if release else "DRAFT"
    path = os.path.join(out_dir, f"{base_name}_{suffix}.xlsx")
    wb.save(path)
    _normalize_xlsx(path)
    return {"path": path, "status": "RELEASE" if release else "DRAFT", "hard_fail": hard_fail,
            "quarantined": len(q), "stamp": status}


def _validation_sheet(wb, checks, run_ts):
    ws = wb.create_sheet("VALIDATION")
    for j, h in enumerate(["check", "tier", "status", "fail_count", "detail", "run_ts"], 1):
        ws.cell(1, j, h).font = _BOLD
    for i, c in enumerate(checks, 2):
        ws.cell(i, 1, c["name"]); ws.cell(i, 2, c["tier"]); ws.cell(i, 3, c["status"])
        ws.cell(i, 4, c["fail"]); ws.cell(i, 5, (c.get("detail") or "")[:200]); ws.cell(i, 6, run_ts)


def _quarantine_sheet(wb, q):
    ws = wb.create_sheet("QUARANTINE")
    for j, h in enumerate(["sheet", "row_id", "field", "reason", "original_value"], 1):
        ws.cell(1, j, h).font = _BOLD
    for i, e in enumerate(q, 2):
        ws.cell(i, 1, e["sheet"]); ws.cell(i, 2, e["row_id"]); ws.cell(i, 3, e["field"])
        ws.cell(i, 4, e["reason"]); ws.cell(i, 5, e.get("original", ""))


def _reliability_sheet(wb, tel, run_ts):
    ws = wb.create_sheet("RELIABILITY")
    for j, h in enumerate(["metric", "value"], 1):
        ws.cell(1, j, h).font = _BOLD
    conf = tel.get("confidence", {})
    rowvals = [("run_ts", run_ts), ("numbers", tel.get("numbers", 0)),
               ("confirmed", tel.get("confirmed", 0)), ("confirmed_rate", tel.get("confirmed_rate", 0)),
               ("quarantined", tel.get("quarantined", 0)), ("quarantine_rate", tel.get("quarantine_rate", 0)),
               ("needs_vision", tel.get("needs_vision", 0)), ("hard_fail", tel.get("hard_fail", 0)),
               ("soft_fail", tel.get("soft_fail", 0)), ("guarded_fail", tel.get("guarded_fail", 0)),
               ("conf_high", conf.get("high", 0)), ("conf_medium", conf.get("medium", 0)),
               ("conf_low", conf.get("low", 0))]
    for i, (k, v) in enumerate(rowvals, 2):
        ws.cell(i, 1, k); ws.cell(i, 2, v)


def _normalize_xlsx(path):
    """Rewrite the xlsx zip with sorted entries + fixed timestamps -> byte-deterministic output."""
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin:
        items = sorted(zin.infolist(), key=lambda i: i.filename)
        payload = [(it.filename, zin.read(it.filename), it.external_attr) for it in items]
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data, attr in payload:
            zi = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = attr
            zout.writestr(zi, data)
    os.replace(tmp, path)


def extract(data: Dict, out_dir: str, grids=None, spans=None, pdf_reader=None, image_reader=None,
            alpha: Optional[float] = None, run_ts: Optional[str] = None, base_name: str = "extraction") -> Dict:
    """End-to-end: reliability pass -> workbook. Deterministic; run_ts is injected (never wall-clock)."""
    clean, report = REL.process(data, grids=grids, spans=spans, pdf_reader=pdf_reader,
                                image_reader=image_reader, alpha=alpha, run_ts=run_ts)
    return build_workbook(clean, report, out_dir, base_name=base_name, alpha=alpha or 0.05)
